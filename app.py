from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import sqlite3
from datetime import datetime
import qrcode
import os

app = Flask(__name__)

# ---------------- DATABASE ----------------
def init_db():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    c.execute('''
    CREATE TABLE IF NOT EXISTS assets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cpu_name TEXT,
        serial_number TEXT,
        status TEXT,
        scan_count INTEGER DEFAULT 0,
        last_updated TEXT
    )
    ''')

    conn.commit()
    conn.close()

init_db()

# ---------------- EXCEL SETUP ----------------
EXCEL_FILE = "assets.xlsx"

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"
    ws.append(["CPU Name", "Serial Number", "Status", "Last Updated"])
    wb.save(EXCEL_FILE)

# ---------------- SAFE EXCEL FUNCTION ----------------
def save_to_excel(cpu_name, serial, status, now):
    # Always ensure file exists
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Assets"
        ws.append(["CPU Name", "Serial Number", "Status", "Last Updated"])
        wb.save(EXCEL_FILE)

    # Load + append safely
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    ws.append([cpu_name, serial, status, now])

    wb.save(EXCEL_FILE)

# ---------------- HOME PAGE ----------------
@app.route('/')
def index():
    return render_template('add.html')

# ---------------- ADD ASSET ----------------
@app.route('/add', methods=['POST'])
def add():
    cpu_name = request.form['cpu_name']
    serial = request.form['serial']
    status = request.form['status']

    now = datetime.now().strftime("%d-%m-%Y %H:%M")

    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    c.execute("""
        INSERT INTO assets (cpu_name, serial_number, status, last_updated)
        VALUES (?, ?, ?, ?)
    """, (cpu_name, serial, status, now))

    asset_id = c.lastrowid

    conn.commit()
    conn.close()

    # ---------------- SAVE TO EXCEL (FIXED) ----------------
    save_to_excel(cpu_name, serial, status, now)

    # ---------------- QR CODE ----------------
    url = f"http://192.168.200.117:5000/asset/{asset_id}"
    qr = qrcode.make(url)

    if not os.path.exists("static"):
        os.makedirs("static")

    qr_path = f"static/qr_{asset_id}.png"
    qr.save(qr_path)

    return f"""
    <h2>Asset Added Successfully!</h2>
    <p>Scan this QR:</p>
    <img src='/{qr_path}' width='200'>
    <br><a href='/'>Go Back</a>
    """

# ---------------- VIEW ASSET ----------------
@app.route('/asset/<int:id>')
def asset(id):
    conn = sqlite3.connect('database.db')
    c = conn.cursor()

    c.execute("SELECT * FROM assets WHERE id=?", (id,))
    data = c.fetchone()

    if not data:
        return "Asset not found"

    new_count = data[4] + 1

    c.execute("UPDATE assets SET scan_count=? WHERE id=?", (new_count, id))
    conn.commit()
    conn.close()

    return render_template('asset.html', data=data, scan=new_count)

# ---------------- RUN SERVER ----------------
if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000, debug=True)
